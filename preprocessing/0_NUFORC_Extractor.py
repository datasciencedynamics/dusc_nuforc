#!/usr/bin/env python3
"""
nuforc_enrich_final.py
----------------------
Adds columns to NUFORC_DATA_*.xlsx by visiting the link in column A for each
row and parsing the sighting detail page.

Confirmed: the report content (bold fields + narrative) IS in the server HTML,
so plain requests + BeautifulSoup works. No browser / Selenium needed. The only
real constraint is request RATE (NUFORC runs Wordfence, which IP-bans you if you
go too fast).

By default it appends exactly three columns after the last column:
    Full_Text      - the full narrative report text
    Color          - the "Color:" bold field (blank on older reports that
                     don't have it; that's expected, not a bug)
    Reported_Full  - the "Reported:" field with full timestamp + timezone,
                     which the existing date-only "Reported" column lacks
Pass --grab-all-fields to also pull Duration, Shape detail, Characteristics, etc.

RESUMABLE: every scraped row is logged to a checkpoint file. If you get banned
or the laptop sleeps, just run the same command again. It skips everything
already done and only fetches what's missing.

SETUP (one time):
    pip install requests beautifulsoup4 openpyxl tqdm

RUN (standalone, uses the defaults below):
    python nuforc_enrich_final.py
    # mac, keep awake:   caffeinate -i python nuforc_enrich_final.py

RUN (via make, args override the defaults):
    python nuforc_enrich_final.py \
        --input-data-file ./data/raw/NUFORC_DATA_04_10_2026.xlsx \
        --output-data-file ./data/raw/NUFORC_DATA_04_10_2026_enriched.xlsx
"""

import argparse
import json
import random
import re
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from tqdm import tqdm

# --------------------------- DEFAULTS (CONFIG) ------------------------------
# These double as the argparse defaults, so the script runs with no args AND
# via make with explicit --flags.
INPUT_FILE = "NUFORC_DATA_04_10_2026.xlsx"
OUTPUT_FILE = "NUFORC_DATA_04_10_2026_enriched.xlsx"
CHECKPOINT = "nuforc_enrich_checkpoint.jsonl"
SHEET_NAME = None  # None = active sheet
LINK_COL = 1  # column A

# Rate control. Sean's 0.5s got him banned; 30s is safe but ~8 days for 20k.
# 5-8s threads the needle for an overnight-ish run. If you get blocked, the run
# stops cleanly and you just bump these up and rerun (it resumes).
MIN_DELAY = 5.0
MAX_DELAY = 8.0

SAVE_EVERY = 25
REQUEST_TIMEOUT = 30
GRAB_ALL_FIELDS = False  # True -> also append every other bold field
TEST_LIMIT = None  # e.g. 10 to try the first 10 rows first, then None

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
# ----------------------------------------------------------------------------

ID_RE = re.compile(r"[?&]id=(\d+)")

# Bold labels -> output column name. Keep the 3 you want first.
CORE_LABELS = {
    "Reported": "Reported_Full",
    "Color": "Color",
}
EXTRA_LABELS = {
    "Occurred": "Occurred_Detail",
    "Duration": "Duration_Detail",
    "No of observers": "Observers_Detail",
    "Location": "Location_Detail",
    "Location details": "Location_details",
    "Shape": "Shape_Detail",
    "Estimated Size": "Estimated_Size",
    "Viewed From": "Viewed_From",
    "Direction from Viewer": "Direction_from_Viewer",
    "Angle of Elevation": "Angle_of_Elevation",
    "Heading": "Heading",
    "Closest Distance": "Closest_Distance",
    "Estimated Speed": "Estimated_Speed",
    "Characteristics": "Characteristics",
}

NAV_FOOTER = {
    "Skip to content",
    "Menu Close",
    "Posts",
    "Data Bank",
    "Map",
    "Gallery",
    "File a UFO Report",
    "Donate",
    "About Us",
    "Toggle website search",
    "TERMS OF SERVICE",
    "PRIVACY POLICY",
}
BLOCK_PHRASES = (
    "Your access to this site has been limited",
    "Exceeded the maximum number of requests",
    "HTTP response code 503",
    "Block Technical Data",
)


def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip() if s is not None else ""


def sighting_id(url):
    m = ID_RE.search(url or "")
    return m.group(1) if m else None


def value_after_b(b_tag):
    """Text after <b>Label:</b> up to the next <br>."""
    out = []
    for sib in b_tag.next_siblings:
        if getattr(sib, "name", None) == "br":
            break
        txt = sib.get_text(" ", strip=True) if hasattr(sib, "get_text") else str(sib)
        txt = clean(txt)
        if txt:
            out.append(txt)
    return clean(" ".join(out))


def parse_page(html, label_map, output_cols):
    soup = BeautifulSoup(html, "html.parser")
    primary = soup.find(id="primary") or soup

    fields = {col: "" for col in output_cols}

    # Capture EVERY bold field's value (not just the ones we output) so the
    # narrative extractor knows all the structured values to exclude.
    field_values = set()  # exact value strings to drop from narrative
    bare_labels = set()  # label texts, e.g. "Reported", to drop label lines
    for b in primary.find_all("b"):
        label = clean(b.get_text(" ", strip=True)).rstrip(":")
        val = value_after_b(b)
        bare_labels.add(label)
        if val:
            field_values.add(val.lower())
        if label in label_map:
            fields[label_map[label]] = val

    # Narrative: flatten <br> to newlines, then keep only witness lines.
    for br in primary.find_all("br"):
        br.replace_with("\n")
    raw = primary.get_text("\n", strip=True)

    lines, seen = [], set()
    for line in (clean(x) for x in raw.split("\n")):
        if not line:
            continue
        if line.startswith("NUFORC UFO Sighting"):
            continue
        if line.startswith("Posted ") or re.match(r"Posted\s+\d{4}-\d{2}-\d{2}", line):
            break
        low = line.lower()
        if low in field_values:  # a structured value
            continue
        if line.rstrip(":") in bare_labels:  # a bare label line
            continue
        if line in NAV_FOOTER or line.startswith("Copyright"):
            continue
        if low in seen:  # dedupe repeated summary
            continue
        seen.add(low)
        lines.append(line)

    fields["Full_Text"] = "\n".join(lines).strip()
    return fields


def is_blocked(text):
    return any(p in text for p in BLOCK_PHRASES)


def load_checkpoint(path):
    done = {}
    p = Path(path)
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
                done[str(rec["id"])] = rec
            except Exception:
                pass
    return done


def append_checkpoint(path, rec):
    with open(path, "a", encoding="utf-8") as fh:
        fh.write(json.dumps(rec, ensure_ascii=False) + "\n")


def parse_args():
    p = argparse.ArgumentParser(
        description="Scrape NUFORC sighting detail pages into the workbook."
    )
    p.add_argument("--input-data-file", default=INPUT_FILE)
    p.add_argument("--output-data-file", default=OUTPUT_FILE)
    p.add_argument("--checkpoint", default=CHECKPOINT)
    p.add_argument("--sheet-name", default=SHEET_NAME)
    p.add_argument("--link-col", type=int, default=LINK_COL)
    p.add_argument("--min-delay", type=float, default=MIN_DELAY)
    p.add_argument("--max-delay", type=float, default=MAX_DELAY)
    p.add_argument("--save-every", type=int, default=SAVE_EVERY)
    p.add_argument("--request-timeout", type=int, default=REQUEST_TIMEOUT)
    p.add_argument("--grab-all-fields", action="store_true", default=GRAB_ALL_FIELDS)
    p.add_argument("--test-limit", type=int, default=TEST_LIMIT)
    p.add_argument("--user-agent", default=USER_AGENT)
    return p.parse_args()


def main():
    args = parse_args()

    label_map = dict(CORE_LABELS)
    if args.grab_all_fields:
        label_map.update(EXTRA_LABELS)
    # narrative first, then mapped labels in insertion order
    output_cols = ["Full_Text"] + list(dict.fromkeys(label_map.values()))

    if not Path(args.input_data_file).exists():
        raise SystemExit(f"Can't find {args.input_data_file}")

    wb = load_workbook(args.input_data_file)
    ws = wb[args.sheet_name] if args.sheet_name else wb.active

    # Append output headers in the first empty columns (preserve everything).
    headers = {
        clean(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)
    }
    for name in output_cols:
        if name not in headers:
            c = ws.max_column + 1
            ws.cell(row=1, column=c, value=name)
            headers[name] = c

    done = load_checkpoint(args.checkpoint)
    print(f"{len(done)} rows already done (resuming).")

    session = requests.Session()
    session.headers.update({"User-Agent": args.user_agent})

    last_row = (
        ws.max_row if args.test_limit is None else min(ws.max_row, 1 + args.test_limit)
    )
    rows = list(range(2, last_row + 1))

    def url_of(r):
        cell = ws.cell(row=r, column=args.link_col)
        if cell.hyperlink:
            return cell.hyperlink.target
        return cell.value if str(cell.value).startswith("http") else None

    # progress bar counts only rows still needing a live fetch
    todo = sum(1 for r in rows if (s := sighting_id(url_of(r))) and s not in done)
    bar = tqdm(total=todo, unit="page", desc="NUFORC", dynamic_ncols=True)

    since_save = 0
    blocked = False
    for r in rows:
        url = url_of(r)
        sid = sighting_id(url)
        if not sid:
            continue

        if sid in done:
            rec = done[sid]
        else:
            try:
                resp = session.get(url, timeout=args.request_timeout)
                if resp.status_code == 503 or is_blocked(resp.text):
                    print(
                        "\nRate-limit/block hit. Saving and stopping cleanly. "
                        "Raise --min-delay/--max-delay and rerun to resume."
                    )
                    blocked = True
                    break
                resp.raise_for_status()
                rec = {"id": sid, **parse_page(resp.text, label_map, output_cols)}
            except Exception as e:
                rec = {
                    "id": sid,
                    **{c: "" for c in output_cols},
                    "Full_Text": f"SCRAPE FAILED: {e}",
                }
            append_checkpoint(args.checkpoint, rec)
            done[sid] = rec
            bar.update(1)
            bar.set_postfix_str(f"id={sid} color={rec.get('Color', '')!r}")
            time.sleep(random.uniform(args.min_delay, args.max_delay))

        for name in output_cols:
            ws.cell(row=r, column=headers[name], value=rec.get(name, ""))

        since_save += 1
        if since_save >= args.save_every:
            wb.save(args.output_data_file)
            since_save = 0

    bar.close()
    wb.save(args.output_data_file)
    print(f"\nSaved {Path(args.output_data_file).resolve()}")
    if blocked:
        print("Stopped early on a block. Rerun the same command to continue.")
    else:
        print("All requested rows processed.")


if __name__ == "__main__":
    main()
