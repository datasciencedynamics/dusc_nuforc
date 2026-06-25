"""
NUFORC Detail Extractor

Reads hyperlinks from an Excel workbook and extracts detail-page fields into Excel.

Install:
pip install beautifulsoup4 requests openpyxl
"""

import re
import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook



# SETTINGS


INPUT_FILE = r"C:\Users\SeanT\Downloads\NUFORC_DATA_04_10_2026.xlsx"
OUTPUT_FILE = r"C:\Users\SeanT\Downloads\NUFORC_DATA_04_10_2026_EXTRACTED.xlsx"

SHEET_NAME = None
LINK_COL = 1

START_ROW = 2
END_ROW = 20  # change to None for full file

REQUEST_DELAY_SECONDS = 30
SAVE_EVERY = 10


DETAIL_FIELDS = [
    "Sighting ID",
    "Detail URL",
    "Occurred Detail",
    "Reported Detail",
    "Duration Detail",
    "No of observers",
    "Location Detail",
    "Location details",
    "Shape Detail",
    "Color",
    "Estimated Size",
    "Viewed From",
    "Direction from Viewer",
    "Angle of Elevation",
    "Heading",
    "Closest Distance",
    "Estimated Speed",
    "Characteristics",
    "Description",
    "Posted",
]


LABEL_MAP = {
    "Occurred": "Occurred Detail",
    "Reported": "Reported Detail",
    "Duration": "Duration Detail",
    "No of observers": "No of observers",
    "Location": "Location Detail",
    "Location details": "Location details",
    "Shape": "Shape Detail",
    "Color": "Color",
    "Estimated Size": "Estimated Size",
    "Viewed From": "Viewed From",
    "Direction from Viewer": "Direction from Viewer",
    "Angle of Elevation": "Angle of Elevation",
    "Heading": "Heading",
    "Closest Distance": "Closest Distance",
    "Estimated Speed": "Estimated Speed",
    "Characteristics": "Characteristics",
}



# HELPER FUNCTIONS


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def get_hyperlink_url(cell):
    if cell.hyperlink:
        return cell.hyperlink.target

    value = clean_text(cell.value)

    if value.startswith("http"):
        return value

    return None


def safe_sleep():
    delay = REQUEST_DELAY_SECONDS + random.uniform(0, 5)
    print(f"Waiting {round(delay, 1)} seconds...")
    time.sleep(delay)


def get_text_after_b_tag(b_tag):
    """
    Extracts value after a label like:
    <b>Color:</b> Orange<br>
    """
    pieces = []

    for sibling in b_tag.next_siblings:
        if getattr(sibling, "name", None) == "br":
            break

        if hasattr(sibling, "get_text"):
            text = sibling.get_text(" ", strip=True)
        else:
            text = str(sibling)

        text = clean_text(text)

        if text:
            pieces.append(text)

    return clean_text(" ".join(pieces))


def parse_nuforc_page(html, url):
    soup = BeautifulSoup(html, "html.parser")

    primary = soup.find(id="primary")
    if primary is None:
        primary = soup

    data = {field: "" for field in DETAIL_FIELDS}
    data["Detail URL"] = url

    title = primary.find("h1")
    if title:
        title_text = clean_text(title.get_text(" ", strip=True))
        match = re.search(r"(\d+)", title_text)
        if match:
            data["Sighting ID"] = match.group(1)

    for b in primary.find_all("b"):
        label = clean_text(b.get_text(" ", strip=True)).replace(":", "")

        if label in LABEL_MAP:
            field_name = LABEL_MAP[label]
            data[field_name] = get_text_after_b_tag(b)

    posted_match = re.search(
        r"Posted\s+(\d{4}-\d{2}-\d{2})",
        primary.get_text(" ", strip=True)
    )

    if posted_match:
        data["Posted"] = posted_match.group(1)

    for br in primary.find_all("br"):
        br.replace_with("\n")

    text = primary.get_text("\n", strip=True)
    lines = [clean_text(line) for line in text.split("\n") if clean_text(line)]

    skip_exact = set(v for v in data.values() if v)
    skip_prefixes = tuple(label + ":" for label in LABEL_MAP.keys())

    description_lines = []

    for line in lines:
        if line.startswith("NUFORC UFO Sighting"):
            continue

        if line.startswith(skip_prefixes):
            continue

        if line.startswith("Posted "):
            break

        if line in skip_exact:
            continue

        if len(line) > 25:
            description_lines.append(line)

    data["Description"] = clean_text(" ".join(description_lines))

    return data


def is_blocked_page(text):
    block_phrases = [
        "Your access to this site has been limited",
        "Exceeded the maximum number of requests per minute",
        "HTTP response code 503",
        "Block Technical Data",
    ]

    return any(phrase in text for phrase in block_phrases)


def scrape_url(session, url, retries=1):
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, timeout=30)

            if response.status_code == 503 or is_blocked_page(response.text):
                print("Rate limit/block detected. Saving and stopping.")
                return "BLOCKED"

            response.raise_for_status()
            return parse_nuforc_page(response.text, url)

        except Exception as e:
            print(f"Attempt {attempt} failed for {url}: {e}")
            time.sleep(30 * attempt)

    failed = {field: "" for field in DETAIL_FIELDS}
    failed["Detail URL"] = url
    failed["Description"] = "SCRAPE FAILED"
    return failed



# MAIN PROGRAM


def main():
    wb = load_workbook(INPUT_FILE)

    if SHEET_NAME:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    existing_headers = {
        clean_text(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
    }

    for field in DETAIL_FIELDS:
        if field not in existing_headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = field
            existing_headers[field] = new_col

    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
        )
    })

    total_rows = ws.max_row
    final_row = total_rows if END_ROW is None else min(END_ROW, total_rows)

    for row in range(START_ROW, final_row + 1):
        link_cell = ws.cell(row=row, column=LINK_COL)
        url = get_hyperlink_url(link_cell)

        if not url:
            print(f"Row {row}: No hyperlink found")
            continue

        print(f"Row {row}/{total_rows}: {url}")

        detail_data = scrape_url(session, url)

        if detail_data == "BLOCKED":
            wb.save(OUTPUT_FILE)
            print(f"Stopped due to rate limit.")
            print(f"Progress saved to: {OUTPUT_FILE}")
            break

        for field in DETAIL_FIELDS:
            col = existing_headers[field]
            ws.cell(row=row, column=col).value = detail_data.get(field, "")

        if row % SAVE_EVERY == 0:
            wb.save(OUTPUT_FILE)
            print(f"Saved checkpoint at row {row}")

        safe_sleep()

    wb.save(OUTPUT_FILE)
    print(f"Done or safely stopped. Saved as: {OUTPUT_FILE}")


main()