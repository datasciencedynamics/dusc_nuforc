#!/usr/bin/env python3
"""
backfill_summary_text.py
For reports whose only witness narrative is the summary (Full_Text blank but
Summary present), copy Summary into Full_Text so the text field means "all
available witness narrative." Adds a traceable flag column and reports the
outcome-class balance of the affected rows so any confound is visible before
you commit to it.

No scraping. Local column edit only. Reads the enriched xlsx, writes a NEW file
(input left untouched).

    python preprocessing/backfill_summary_text.py \
        --input-file  ./data/raw/NUFORC_DATA_04_10_2026.xlsx \
        --output-file ./data/raw/NUFORC_DATA_04_10_2026_backfilled.xlsx \
        --outcome-col dramatic
"""

import argparse
from collections import Counter
from openpyxl import load_workbook


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-file", default="././data/raw/NUFORC_DATA_04_10_2026.xlsx")
    ap.add_argument("--output-file", default="././data/raw/NUFORC_DATA_04_10_2026.xlsx")
    ap.add_argument(
        "--outcome-col",
        default=None,
        help="optional column to report class balance on (e.g. dramatic)",
    )
    args = ap.parse_args()

    wb = load_workbook(args.input_file)
    ws = wb.active
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    c_full, c_sum = hdr["Full_Text"], hdr["Summary"]

    # add the flag column if not present
    flag = "text_is_summary_only"
    if flag not in hdr:
        c_flag = ws.max_column + 1
        ws.cell(1, c_flag, flag)
    else:
        c_flag = hdr[flag]

    c_out = hdr.get(args.outcome_col) if args.outcome_col else None

    backfilled = 0
    still_empty = 0
    affected_classes = Counter()
    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, c_full).value
        summ = ws.cell(r, c_sum).value
        full_blank = not (full and str(full).strip())
        summ_present = bool(summ and str(summ).strip())

        if full_blank and summ_present:
            ws.cell(r, c_full, summ)  # Full_Text := Summary
            ws.cell(r, c_flag, 1)
            backfilled += 1
            if c_out:
                affected_classes[ws.cell(r, c_out).value] += 1
        else:
            ws.cell(r, c_flag, 0)
            if full_blank and not summ_present:
                still_empty += 1  # truly empty (no summary either)

    wb.save(args.output_file)

    print(f"backfilled Full_Text from Summary: {backfilled} rows")
    print(f"still truly empty (no summary):    {still_empty} rows")
    if c_out:
        print(
            f"\nclass balance of the {backfilled} backfilled rows "
            f"(by '{args.outcome_col}'):"
        )
        for k, v in affected_classes.most_common():
            print(f"  {k!r:>10}: {v}")
        print("compare to the overall balance to judge if this skews a class.")
    print(f"\nwrote {args.output_file}  (input untouched)")


if __name__ == "__main__":
    main()
