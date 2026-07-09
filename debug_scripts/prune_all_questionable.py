#!/usr/bin/env python3
"""
prune_all_questionable.py
One-pass cleanup: drop every questionable row from the checkpoint so a single
re-scrape (with the fixed parser) fixes all of them at once instead of chasing
one category at a time.

Flags a row for re-scrape if its Full_Text is:
  - a SCRAPE FAILED marker (transient fetch error), OR
  - empty, OR
  - exactly equal to Summary (the truncation fingerprint)

Backs up the checkpoint first and writes the flagged ids to a record file.
Read-only on the xlsx; only the checkpoint is rewritten.
"""

import re
import json
import shutil
import argparse
from openpyxl import load_workbook

norm = lambda s: re.sub(r"\s+", " ", str(s or "")).strip().lower()
ID_RE = re.compile(r"[?&]id=(\d+)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--enriched-file", default="././data/raw/NUFORC_DATA_04_10_2026.xlsx"
    )
    ap.add_argument(
        "--checkpoint", default="./data/raw/nuforc_enrich_checkpoint_TEST.jsonl"
    )
    ap.add_argument("--link-col", type=int, default=1)
    args = ap.parse_args()

    ws = load_workbook(args.enriched_file).active
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    c_full, c_sum = hdr["Full_Text"], hdr["Summary"]

    reasons = {"failed": set(), "empty": set(), "eq_summary": set()}
    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, c_full).value
        summ = ws.cell(r, c_sum).value
        link = ws.cell(r, args.link_col)
        url = link.hyperlink.target if link.hyperlink else link.value
        m = ID_RE.search(url or "")
        if not m:
            continue
        sid = m.group(1)

        s = str(full or "")
        if s.startswith("SCRAPE FAILED"):
            reasons["failed"].add(sid)
        elif not s.strip():
            reasons["empty"].add(sid)
        elif norm(full) == norm(summ):
            reasons["eq_summary"].add(sid)

    bad = reasons["failed"] | reasons["empty"] | reasons["eq_summary"]

    print(f"flagged {len(bad)} rows for re-scrape:")
    print(f"  SCRAPE FAILED       {len(reasons['failed'])}")
    print(f"  empty               {len(reasons['empty'])}")
    print(f"  Full_Text==Summary  {len(reasons['eq_summary'])}")

    # record ids (with reason) before touching anything
    with open("./data/raw/prune_all_ids.txt", "w") as f:
        for reason, ids in reasons.items():
            for sid in sorted(ids):
                f.write(f"{sid}\t{reason}\n")

    # back up the checkpoint, then drop the flagged ids
    shutil.copy(args.checkpoint, args.checkpoint + ".safe")
    lines = [json.loads(l) for l in open(args.checkpoint)]
    keep = [l for l in lines if str(l["id"]) not in bad]
    with open(args.checkpoint, "w") as f:
        for l in keep:
            f.write(json.dumps(l, ensure_ascii=False) + "\n")

    print(f"\ncheckpoint backed up to {args.checkpoint}.safe")
    print(f"dropped {len(lines) - len(keep)} checkpoint lines, kept {len(keep)}")
    print("now run: make scrape_nuforc_details   (re-fetches only the flagged ids)")


if __name__ == "__main__":
    main()
