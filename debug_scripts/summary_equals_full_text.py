#!/usr/bin/env python3
"""
export_summary_equals_fulltext.py
Isolate the rows where Full_Text is empty or equals Summary (the audit's
"bug" bucket) into a separate workbook for manual review. Read-only on the
source; writes a new file. Column A stays a clickable link to the report.
"""

import re
import argparse
from openpyxl import load_workbook, Workbook

norm = lambda s: re.sub(r"\s+", " ", str(s or "")).strip().lower()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--enriched-file", default="././data/raw/NUFORC_DATA_04_10_2026.xlsx"
    )
    ap.add_argument(
        "--output-file", default="././data/raw/NUFORC_review_summary_eq_full_text.xlsx"
    )
    ap.add_argument("--link-col", type=int, default=1)
    args = ap.parse_args()

    ws = load_workbook(args.enriched_file).active
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    c_sum = hdr["Summary"]
    c_full = hdr["Full_Text"]

    # Columns to carry into the review file (only those that exist).
    want = [
        "Summary",
        "Full_Text",
        "Color",
        "Reported_Full",
        "City",
        "State",
        "Shape",
        "Occurred",
        "Reported",
    ]
    keep_cols = [(name, hdr[name]) for name in want if name in hdr]

    out = Workbook()
    ows = out.active
    ows.title = "review"
    # header row: Link + reason + chosen columns
    ows.append(["Link", "reason"] + [name for name, _ in keep_cols])

    n = 0
    for r in range(2, ws.max_row + 1):
        full = ws.cell(r, c_full).value
        summ = ws.cell(r, c_sum).value
        is_empty = not (full and str(full).strip())
        is_eq = (not is_empty) and norm(full) == norm(summ)
        if not (is_empty or is_eq):
            continue

        link_cell = ws.cell(r, args.link_col)
        url = link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value
        reason = "empty" if is_empty else "full==summary"

        row_vals = [url, reason] + [ws.cell(r, col).value for _, col in keep_cols]
        ows.append(row_vals)
        n += 1
        # make column A an actual clickable hyperlink
        cell = ows.cell(row=n + 1, column=1)
        if url and str(url).startswith("http"):
            cell.hyperlink = url
            cell.style = "Hyperlink"

    out.save(args.output_file)
    print(f"wrote {n} rows to {args.output_file}")


if __name__ == "__main__":
    main()
