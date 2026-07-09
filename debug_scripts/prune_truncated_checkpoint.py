import re, json
from openpyxl import load_workbook

norm = lambda s: re.sub(r"\s+", " ", str(s or "")).strip().lower()
ID_RE = re.compile(r"[?&]id=(\d+)")

ws = load_workbook("././data/raw/NUFORC_DATA_04_10_2026.xlsx").active
hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

bad = set()
for r in range(2, ws.max_row + 1):
    full = ws.cell(row=r, column=hdr["Full_Text"]).value
    summ = ws.cell(row=r, column=hdr["Summary"]).value
    is_empty = not (full and str(full).strip())
    if is_empty or norm(full) == norm(summ):
        link = ws.cell(row=r, column=1)
        url = link.hyperlink.target if link.hyperlink else link.value
        m = ID_RE.search(url or "")
        if m:
            bad.add(m.group(1))

# record exactly which ids were flagged, for traceability
open("././data/raw/pruned_ids.txt", "w").write("\n".join(sorted(bad)))

path = "././data/raw/nuforc_enrich_checkpoint_TEST.jsonl"
lines = [json.loads(l) for l in open(path)]
keep = [l for l in lines if str(l["id"]) not in bad]
with open(path, "w") as f:
    for l in keep:
        f.write(json.dumps(l, ensure_ascii=False) + "\n")

print(
    f"flagged {len(bad)} bad ids; checkpoint dropped {len(lines)-len(keep)}, kept {len(keep)}"
)
