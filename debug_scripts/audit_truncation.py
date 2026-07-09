#!/usr/bin/env python3
"""
audit_truncation.py
Buckets every row by several truncation signals so we can see the real numbers
instead of guessing with a char threshold. Read-only; changes nothing.

Signals:
  empty        - Full_Text blank (body + summary both lost)
  eq_summary   - Full_Text == Summary  (THE bug fingerprint: body dropped)
  zero_nl      - Full_Text has no newline (single line: truncated OR genuinely
                 short one-liner; broader, catches more but less precise)
  multiline    - Full_Text has >=1 newline (almost certainly fine)

Recommended re-scrape set = empty OR eq_summary.
zero_nl is reported too so you can compare and widen the net if you want.

Run from the folder with the enriched xlsx:
    python audit_truncation.py
"""

import re
from openpyxl import load_workbook

ENRICHED = "././data/raw/NUFORC_DATA_04_10_2026.xlsx"
norm = lambda s: re.sub(r"\s+", " ", str(s or "")).strip().lower()

wb = load_workbook(ENRICHED)
ws = wb.active
hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
c_sum, c_full = hdr["Summary"], hdr["Full_Text"]

total = empty = eq_summary = zero_nl = multiline = 0
rescrape = 0  # empty OR eq_summary
zero_nl_not_summary = 0  # single-line but NOT equal to summary (genuine shorts)

for r in range(2, ws.max_row + 1):
    full = ws.cell(row=r, column=c_full).value
    summ = ws.cell(row=r, column=c_sum).value
    total += 1

    is_empty = not (full and str(full).strip())
    is_eq = (not is_empty) and norm(full) == norm(summ)
    is_zero_nl = (not is_empty) and ("\n" not in str(full))
    is_multi = (not is_empty) and ("\n" in str(full))

    if is_empty:
        empty += 1
    if is_eq:
        eq_summary += 1
    if is_zero_nl:
        zero_nl += 1
        if not is_eq:
            zero_nl_not_summary += 1
    if is_multi:
        multiline += 1
    if is_empty or is_eq:
        rescrape += 1

print(f"total rows                         {total}")
print(f"  empty Full_Text                  {empty}")
print(f"  Full_Text == Summary (the bug)   {eq_summary}")
print(f"  zero-newline (single line)       {zero_nl}")
print(
    f"     of which NOT == summary       {zero_nl_not_summary}  (likely genuine one-liners)"
)
print(f"  multi-line (has newline)         {multiline}")
print("-" * 45)
print(f"RECOMMENDED re-scrape (empty|==sum) {rescrape}")
print(f"WIDER net if paranoid (empty|zeroNL) {empty + zero_nl}")
