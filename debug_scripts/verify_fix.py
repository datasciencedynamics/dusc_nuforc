#!/usr/bin/env python3
"""
verify_fix.py
Confirms the narrative fix on REAL pages before any mass re-scrape.

Truncation detector: a row is truly truncated when its Full_Text equals its
Summary (that's the exact symptom: full text == summarized text). Genuinely
short reports have real body lines and won't match, so they're left alone.

Re-fetches a small sample of the truncated rows with the FIXED parser and prints
old length vs new length. If new >> old on those rows, the fix is good.

Run from the folder with the enriched xlsx + nuforc_enrich_final.py, venv active:
    python verify_fix.py
Then upload verify_out.txt.
"""

import re
import time
import requests
from openpyxl import load_workbook
from preprocessing.step_00_NUFORC_Extractor import parse_page, CORE_LABELS

ENRICHED = "././data/raw/NUFORC_DATA_04_10_2026.xlsx"
HEAD = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
}
SAMPLE = 10

label_map = dict(CORE_LABELS)
output_cols = ["Full_Text"] + list(dict.fromkeys(label_map.values()))
norm = lambda s: re.sub(r"\s+", " ", str(s or "")).strip().lower()
ID_RE = re.compile(r"[?&]id=(\d+)")

out = open("././data/raw/verify_out.txt", "w", encoding="utf-8")


def log(*a):
    s = " ".join(str(x) for x in a)
    print(s)
    out.write(s + "\n")


wb = load_workbook(ENRICHED)
ws = wb.active
hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
c_sum, c_full = hdr.get("Summary"), hdr.get("Full_Text")

truncated = []
for r in range(2, ws.max_row + 1):
    full = ws.cell(row=r, column=c_full).value
    summ = ws.cell(row=r, column=c_sum).value
    if full and norm(full) == norm(summ):
        link = ws.cell(row=r, column=1)
        url = link.hyperlink.target if link.hyperlink else link.value
        m = ID_RE.search(url or "")
        if m:
            truncated.append((m.group(1), full))

log(f"{len(truncated)} truly-truncated rows (Full_Text == Summary)\n")
log(f"Re-fetching first {SAMPLE} with the FIXED parser:\n")
for sid, old in truncated[:SAMPLE]:
    try:
        html = requests.get(
            f"https://nuforc.org/sighting/?id={sid}", headers=HEAD, timeout=30
        ).text
        new = parse_page(html, label_map, output_cols)["Full_Text"]
    except Exception as e:
        log(f"id={sid}  ERROR {e}")
        continue
    verdict = "RECOVERED" if len(new) > len(old) + 20 else "still short"
    log(f"id={sid}  old_len={len(old)}  new_len={len(new)}  -> {verdict}")
    log("   OLD (truncated) text:")
    log("      " + (old or "(empty)"))
    log("   NEW (full) text:")
    log(new or "(empty)")
    log("-" * 70)
    log("")
    time.sleep(3)

out.close()
print("wrote verify_out.txt")
